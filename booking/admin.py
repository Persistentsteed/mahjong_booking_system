# booking/admin.py

from django.contrib import admin
from django.db.models import Count, Q
from django.utils import timezone
from django.conf import settings 
from django import forms
from django.contrib.admin.helpers import ActionForm
from django.shortcuts import redirect
from django.urls import reverse
from collections import defaultdict
import datetime
# 从 accounts.models 导入 CustomUser（确保路径正确）
from accounts.models import CustomUser 
from .models import Store, MahjongTable, Booking

# 新增导入：处理 HTTP 响应和 Excel 文件
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

@admin.register(Store)
class StoreAdmin(admin.ModelAdmin):
    """
    门店模型后台管理
    """
    list_display = ('name', 'address')

@admin.register(MahjongTable)
class MahjongTableAdmin(admin.ModelAdmin):
    """
    麻将桌模型后台管理
    """
    list_display = ('store', 'table_number', 'alias', 'get_current_status')
    ordering = ('store__name', 'table_number')
    list_filter = ('store',)
    actions = ['create_walk_in_booking'] # 管理员操作：创建散客对局

    def create_walk_in_booking(self, request, queryset):
        """
        Admin Action: 为选中的单个牌桌创建散客对局
        """
        if queryset.count() != 1:
            # ★★★ 修复：将 _self_ 改为 self ★★★
            self.message_user(request, "请一次只选择一个牌桌进行此操作。", level='WARNING')
            return

        table = queryset.first()
        now = timezone.now()

        # 检查该桌当前是否已被占用
        if Booking.objects.filter(table=table, status='CONFIRMED', end_time__gt=now).exists():
            # ★★★ 修复：将 _self_ 改为 self ★★★
            self.message_user(request, f"{table} 当前已被占用，无法创建新的对局。", level='ERROR')
            return

        # 创建一个默认的散客对局
        Booking.objects.create(
            creator=request.user,  # 使用当前管理员作为 creator
            store=table.store,
            table=table,
            start_time=now,
            end_time=now + datetime.timedelta(hours=3),
            num_games=4,           # 默认4个半庄
            status='CONFIRMED',    # 直接设置为已成行
            # end_time 会在 Booking 模型的 save 方法中根据 num_games 自动计算
        )
        # ★★★ 修复：将 _self_ 改为 self ★★★
        self.message_user(request, f"已为 {table} 成功创建一个为期3小时的散客对局。")

    create_walk_in_booking.short_description = "为选中牌桌创建散客对局 (占用3小时)"

    def get_current_status(self, obj):
        """
        辅助方法: 获取牌桌的当前占用状态
        """
        now = timezone.now()
        if Booking.objects.filter(table=obj, status='CONFIRMED', end_time__gt=now).exists():
            return "占用中"
        return "空闲"
    get_current_status.short_description = "当前状态"

class BookingExportActionForm(ActionForm):
    start_date = forms.DateField(
        required=False,
        label="开始日期",
        widget=forms.DateInput(attrs={"type": "date"})
    )
    end_date = forms.DateField(
        required=False,
        label="结束日期",
        widget=forms.DateInput(attrs={"type": "date"})
    )


class BookingStageFilter(admin.SimpleListFilter):
    title = "对局阶段"
    parameter_name = "booking_stage"

    def lookups(self, request, model_admin):
        return (
            ('upcoming', '预约对局'),
            ('ongoing', '进行中对局'),
            ('finished', '已完成对局'),
        )

    def queryset(self, request, queryset):
        now = timezone.now()
        if self.value() == 'upcoming':
            return queryset.filter(start_time__gt=now, status__in=['PENDING', 'CONFIRMED'])
        if self.value() == 'ongoing':
            return queryset.filter(
                start_time__lte=now,
                end_time__gte=now,
                status__in=['CONFIRMED']
            )
        if self.value() == 'finished':
            return queryset.filter(
                Q(end_time__lt=now) | Q(status__in=['COMPLETED', 'CANCELED', 'EXPIRED'])
            )
        return queryset


@admin.register(Booking)
class BookingAdmin(admin.ModelAdmin):
    """
    对局预约模型后台管理
    """
    list_display = ('creator', 'store', 'start_time', 'end_time', 'num_games', 'status', 'get_participant_count', 'table')
    list_filter = ('status', BookingStageFilter, 'store', 'start_time')
    search_fields = ('creator__username', 'creator__display_name', 'store__name')
    autocomplete_fields = ['creator', 'participants'] 
    action_form = BookingExportActionForm
    
    # --- 修复 1: 确保参与者、半庄数和结束时间在自定义表单中是非必填项 ---
    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if 'participants' in form.base_fields:
            form.base_fields['participants'].required = False
        
        if 'num_games' in form.base_fields:
            form.base_fields['num_games'].required = False
        if 'end_time' in form.base_fields:
            form.base_fields['end_time'].required = False
            
        return form
    
    # --- 辅助方法：计算参与人数并显示 (与修复无关，保持不变) ---
    def get_participant_count(self, obj):     
        return obj.participants.count()     
    get_participant_count.short_description = '参与人数'       
  
    # --- 管理员 Actions ---
    change_form_template = "admin/booking/booking/change_form.html"
    actions = ['confirm_selected_bookings', 'export_bookings_to_xlsx', 'export_schedule_to_xlsx'] # 在这里添加新的 Action

    def changeform_view(self, request, object_id=None, form_url='', extra_context=None):
        extra_context = extra_context or {}
        tables = MahjongTable.objects.select_related('store').order_by('store__name', 'table_number')
        extra_context['tables_data'] = [
            {
                'id': table.id,
                'store_id': table.store_id,
                'label': table.display_label()
            }
            for table in tables
        ]
        return super().changeform_view(request, object_id, form_url, extra_context)

    def confirm_selected_bookings(self, request, queryset):       
        valid_bookings = queryset.filter(status='PENDING').annotate(p_count=Count('participants')).filter(p_count=4)       
              
        if valid_bookings:       
            updated_count = valid_bookings.update(status='CONFIRMED')       
            # ★★★ 修复：将 _self_ 改为 self ★★★
            self.message_user(request, f"{updated_count} 个满足条件的预约已成功标记为 '已成行'。", level='SUCCESS')     
        else:       
            # ★★★ 修复：将 _self_ 改为 self ★★★
            self.message_user(request, "选中的预约中没有满足成行条件（等待中且满员）的。", level='WARNING')       
          
    confirm_selected_bookings.short_description = "将选中项标记为 '已成行' (仅限满员)"   

    def _filter_queryset_by_dates(self, request, queryset):
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
        start_dt = None
        end_dt = None
        if start_date_str:
            try:
                start_dt = datetime.datetime.strptime(start_date_str, "%Y-%m-%d")
                start_dt = timezone.make_aware(datetime.datetime.combine(start_dt.date(), datetime.time.min))
            except ValueError:
                self.message_user(request, "开始日期格式错误。", level='ERROR')
        if end_date_str:
            try:
                end_dt = datetime.datetime.strptime(end_date_str, "%Y-%m-%d")
                end_dt = timezone.make_aware(datetime.datetime.combine(end_dt.date(), datetime.time.max))
            except ValueError:
                self.message_user(request, "结束日期格式错误。", level='ERROR')
        if start_dt:
            queryset = queryset.filter(start_time__gte=start_dt)
        if end_dt:
            queryset = queryset.filter(start_time__lte=end_dt)
        return queryset, start_dt, end_dt

    def export_bookings_to_xlsx(self, request, queryset):   
        """   
        Admin Action: 导出选中的对局记录为 XLSX 文件  
        """   
        queryset, start_dt, end_dt = self._filter_queryset_by_dates(request, queryset)
        response = HttpResponse(   
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'   
        )   
        response['Content-Disposition'] = 'attachment; filename="mahjong_bookings_export.xlsx"'   
  
        workbook = Workbook()   
        worksheet = workbook.active  
        worksheet.title = "对局记录"   
  
        # 定义表头  
        columns = [   
            ("ID", 5),   
            ("发起人", 15),   
            ("参与者", 30), # 参与者可能较多，宽度大一些  
            ("门店", 15),   
            ("牌桌", 10),   
            ("半庄数", 8),   
            ("开始时间", 20),   
            ("结束时间", 20),   
            ("状态", 10),   
            ("创建时间", 20),   
        ]   
          
        header_font = Font(name='Calibri', bold=True)   
        thin_border = Border(left=Side(style='thin'),   
                             right=Side(style='thin'),   
                             top=Side(style='thin'),   
                             bottom=Side(style='thin'))   
        align_center = Alignment(horizontal="center", vertical="center")   
        align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)   
  
  
        # 写入表头  
        for col_idx, (header_text, width) in enumerate(columns, 1):   
            cell = worksheet.cell(row=1, column=col_idx, value=header_text)   
            cell.font = header_font  
            cell.alignment = align_center  
            cell.border = thin_border  
            worksheet.column_dimensions[get_column_letter(col_idx)].width = width  
          
        row_num = 2  
        for booking in queryset.order_by('start_time'):   
            participants_str = ", ".join([p.display_name or p.username for p in booking.participants.all()])   
              
            # 确保时间以本地时区显示  
            start_time_local = timezone.localtime(booking.start_time) if booking.start_time else ""   
            end_time_local = timezone.localtime(booking.end_time) if booking.end_time else ""   
            created_at_local = timezone.localtime(booking.created_at) if booking.created_at else ""   
  
            # 填充数据  
            worksheet.cell(row=row_num, column=1, value=booking.id).border = thin_border  
            worksheet.cell(row=row_num, column=2, value=booking.creator.display_name or booking.creator.username).border = thin_border  
            worksheet.cell(row=row_num, column=3, value=participants_str).alignment = align_left # 参与者可能多，允许换行  
            worksheet.cell(row=row_num, column=3).border = thin_border  
            worksheet.cell(row=row_num, column=4, value=booking.store.name).border = thin_border  
            worksheet.cell(row=row_num, column=5, value=booking.table.table_number if booking.table else "未分配").border = thin_border  
            worksheet.cell(row=row_num, column=6, value=booking.num_games if booking.num_games is not None else "-").border = thin_border
            worksheet.cell(row=row_num, column=7, value=start_time_local.strftime('%Y-%m-%d %H:%M') if start_time_local else "").border = thin_border
            worksheet.cell(row=row_num, column=8, value=end_time_local.strftime('%Y-%m-%d %H:%M') if end_time_local else "").border = thin_border  
            worksheet.cell(row=row_num, column=9, value=booking.get_status_display()).border = thin_border  
            worksheet.cell(row=row_num, column=10, value=created_at_local.strftime('%Y-%m-%d %H:%M') if created_at_local else "").border = thin_border
              
            row_num += 1  
  
        workbook.save(response) # 将工作簿保存到响应中  
        return response  
  
    export_bookings_to_xlsx.short_description = "导出选中的对局记录为 XLSX"   

    def export_schedule_to_xlsx(self, request, queryset):
        queryset, start_dt, end_dt = self._filter_queryset_by_dates(request, queryset)
        if not start_dt:
            self.message_user(request, "请至少选择一个开始日期以导出对局记录。", level='ERROR')
            return
        if not end_dt:
            end_dt = start_dt
        if end_dt < start_dt:
            start_dt, end_dt = end_dt, start_dt

        bookings = queryset.select_related('store', 'table', 'creator') \
                           .prefetch_related('participants', 'store__tables') \
                           .order_by('start_time')
        if not bookings.exists():
            self.message_user(request, "选定范围内没有预约记录。", level='WARNING')
            return

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="mahjong_schedule.xlsx"'

        workbook = Workbook()
        workbook.remove(workbook.active)

        local_tz = timezone.get_current_timezone()
        start_day = timezone.localtime(start_dt, local_tz).date()
        end_day = timezone.localtime(end_dt, local_tz).date()

        bookings_list = list(bookings)
        store_cache = {b.store_id: b.store for b in bookings_list}

        current_day = start_day
        while current_day <= end_day:
            day_start = datetime.datetime.combine(current_day, datetime.time.min)
            day_start = timezone.make_aware(day_start, local_tz)
            day_end = day_start + datetime.timedelta(days=1)

            for store_id, store in store_cache.items():
                day_store_bookings = [
                    b for b in bookings_list
                    if b.store_id == store_id
                    and timezone.localtime(b.end_time, local_tz) > day_start
                    and timezone.localtime(b.start_time, local_tz) < day_end
                ]
                if not day_store_bookings:
                    continue

                sheet_name = f"{current_day.strftime('%m%d')}-{store.name}"[:31]
                sheet = workbook.create_sheet(title=sheet_name)
                self._build_schedule_sheet(sheet, day_start, store, day_store_bookings, local_tz)

            current_day += datetime.timedelta(days=1)

        workbook.save(response)
        return response

    export_schedule_to_xlsx.short_description = "导出对局记录（按日期范围）"

    def _build_schedule_sheet(self, sheet, day_start, store, bookings, local_tz):
        sheet.cell(row=1, column=1, value="表号")
        sheet.cell(row=2, column=1, value="时间")

        block_headers = ["起止时间", "半庄数", "参与者1", "参与者2", "参与者3", "参与者4"]
        tables = list(store.tables.all().order_by('table_number'))

        table_booking_map = defaultdict(list)
        unassigned = []
        for booking in bookings:
            if booking.table_id:
                table_booking_map[booking.table_id].append(booking)
            else:
                unassigned.append(booking)

        table_blocks = [(t.table_number, t.id) for t in tables]
        if unassigned:
            table_blocks.append(("未分配", None))

        width = len(block_headers)
        for idx, (table_name, table_id) in enumerate(table_blocks):
            base_col = 2 + idx * width
            sheet.merge_cells(start_row=1, start_column=base_col, end_row=1, end_column=base_col + width - 1)
            cell = sheet.cell(row=1, column=base_col, value=str(table_name))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)

            for offset, header in enumerate(block_headers):
                header_cell = sheet.cell(row=2, column=base_col + offset, value=header)
                header_cell.font = Font(bold=True)
                header_cell.alignment = Alignment(horizontal="center", vertical="center")
                sheet.column_dimensions[get_column_letter(base_col + offset)].width = 16 if offset == 0 else 14

            relevant = table_booking_map[table_id] if table_id else unassigned
            self._fill_table_block(sheet, base_col, relevant, day_start, local_tz)

        for hour in range(24):
            row = 3 + hour
            slot_start = day_start + datetime.timedelta(hours=hour)
            slot_end = slot_start + datetime.timedelta(hours=1)
            sheet.cell(row=row, column=1, value=f"{slot_start.strftime('%H:%M')} - {slot_end.strftime('%H:%M')}")

    def _fill_table_block(self, sheet, base_col, bookings, day_start, local_tz):
        def append_cell(cell, text):
            if not text:
                return
            if cell.value:
                cell.value = f"{cell.value}\n{text}"
            else:
                cell.value = text
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        for booking in bookings:
            start_local = timezone.localtime(booking.start_time, local_tz)
            end_local = timezone.localtime(booking.end_time, local_tz)
            hour_offset = int(max(0, min(23, (start_local - day_start).total_seconds() // 3600)))
            row = 3 + hour_offset

            append_cell(sheet.cell(row=row, column=base_col), f"{start_local.strftime('%H:%M')} - {end_local.strftime('%H:%M')}")
            append_cell(sheet.cell(row=row, column=base_col + 1), str(booking.num_games or ""))

            participants = list(booking.participants.all())
            for idx in range(4):
                name = ""
                if idx < len(participants):
                    participant = participants[idx]
                    name = participant.display_name or participant.username
                append_cell(sheet.cell(row=row, column=base_col + 2 + idx), name)

    def response_change(self, request, obj):
        if "_duplicate_and_edit" in request.POST:
            original_participants = list(obj.participants.all())
            obj.pk = None
            obj.created_at = timezone.now()
            obj.save()
            obj.participants.set(original_participants)
            self.message_user(
                request,
                "已复制当前对局，请更改对局时间、半庄数及对局者后保存。",
                level='SUCCESS'
            )
            change_url = reverse("admin:booking_booking_change", args=[obj.pk])
            return redirect(change_url)
        return super().response_change(request, obj)
  
    # --- 核心改动 1: 动态显示和编辑字段 (与修复无关，保持不变) ---   
    def get_fieldsets(self, request, obj=None):   
        fieldsets = []   
        if obj is None:   
            fieldsets.append(('基本信息', {   
                'fields': ('creator', 'store', 'status', 'start_time', 'end_time', 'num_games', 'table')   
            }))   
            fieldsets.append(('参与者', {'fields': ('participants',)}))   
        else:   
            base_info_fields = ['creator', 'store', 'status', 'start_time', 'end_time', 'num_games', 'table']   
            fieldsets.append(('基本信息', {'fields': tuple(base_info_fields)}))   
            fieldsets.append(('参与者', {'fields': ('participants',)})) # 参与者总是需要显示
        return fieldsets  
  
    # --- 修复 2: 优化 get_readonly_fields 逻辑 (与修复无关，保持不变) ---   
    def get_readonly_fields(self, request, obj=None):   
        read_only_fields = ['created_at']   
        if obj:   
            read_only_fields.extend(['creator', 'start_time'])   
        return tuple(read_only_fields)   
  
    # --- 核心改动 2: 筛选出可分配的牌桌 (与修复无关，保持不变) ---   
    def formfield_for_foreignkey(self, db_field, request, **kwargs):   
        if db_field.name == "table":   
            obj_id = request.resolver_match.kwargs.get('object_id')   
            if obj_id:   
                booking = self.get_object(request, obj_id)   
                if booking:   
                    conflicting_bookings = Booking.objects.filter(   
                        store=booking.store,   
                        table__isnull=False,   
                        start_time__lt=booking.end_time,   
                        end_time__gt=booking.start_time  
                    ).exclude(pk=booking.pk).exclude(status__in=['CANCELED', 'EXPIRED'])   
                    occupied_table_ids = conflicting_bookings.values_list('table_id', flat=True)   
                    kwargs["queryset"] = MahjongTable.objects.filter(   
                        store=booking.store  
                    ).exclude(id__in=occupied_table_ids)   
            else:   
                kwargs["queryset"] = MahjongTable.objects.all()   
        return super().formfield_for_foreignkey(db_field, request, **kwargs)       
        
@admin.register(CustomUser)   
class CustomUserAdmin(admin.ModelAdmin):   
    list_display = ('username', 'display_name', 'is_staff', 'is_active')   
    search_fields = ('username', 'display_name')
